import streamlit as st
import pandas as pd
from datetime import datetime, date, timedelta
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from io import BytesIO
import urllib.parse
import json
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
import os
import webbrowser
import fitz
import win32com.client


# --- Page Configuration ---
st.set_page_config(
    page_title="Service & Estimate Dashboard",
    page_icon="🚚",
    layout="wide",
)

# --- Constants for Google Sheets ---
GSHEET_NAME = "Estimate form"
WORKSHEET_INDEX = 1 # Main data sheet
CREDS_FILE = "Credentials.json"
ARCHIVE_SHEET_NAME = "DailyReportArchive"
EOD_SUMMARY_ARCHIVE_SHEET_NAME = "EODSummaryArchive" # New sheet for EOD summary archive

ARCHIVE_SHEET_HEADERS = ["Report Date", "Needs Estimate Creation", "Needs Shipping", "Needs Reminder"]
EOD_ARCHIVE_SHEET_HEADERS = ["Report Date", "Estimate Task Summary", "Reminder Task Summary", "Shipping Task Summary", "AdHoc Shipped Today"]


EXPECTED_COLUMN_ORDER = [
    "RMA", "SPC Code", "Part Number", "S/N", "Description",
    "Fault Comments", "Resolution Comments", "Sender",
    "Estimate Complete Time", "Estimate Complete",
    "Estimate Approved", "Estimate Approved Time",
    "Estimate Sent To Email", "Estimate Sent Time",
    "Reminder Completed", "Reminder Completed Time", "Reminder Contact Method",
    "QA Approved", "QA Approved Time",
    "Shipped", "Shipped Time"
]
ALL_STATUS_COLUMNS = ["Estimate Complete", "Estimate Approved", "Reminder Completed", "QA Approved", "Shipped"]
ALL_TIME_COLUMNS = [col for col in EXPECTED_COLUMN_ORDER if "Time" in col]


# --- Constants for Business Central Link ---
BC_BASE_URL = "https://businesscentral.dynamics.com/7bcfb5b0-27a1-4e18-99d8-ca66570addd8/Production"
BC_COMPANY = "PROD"
BC_PAGE_ID = "70001"
BC_RMA_FIELD_NAME = "No."
BC_LINK_COL_NAME = "View in BC"

# --- Helper Functions ---
@st.cache_data(ttl=300)
def load_data_from_google_sheet(
    sheet_name=GSHEET_NAME,
    worksheet_index=WORKSHEET_INDEX,
    creds_file=CREDS_FILE
):
    """Loads data from the specified Google Sheet."""
    try:
        scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/spreadsheets",
                 "https://www.googleapis.com/auth/drive.file", "https://www.googleapis.com/auth/drive"]
        creds = ServiceAccountCredentials.from_json_keyfile_name(creds_file, scope)
        client = gspread.authorize(creds)
        spreadsheet = client.open(sheet_name)
        worksheet = spreadsheet.get_worksheet(worksheet_index)

        all_values = worksheet.get_all_values()

        if not all_values:
            return pd.DataFrame(columns=EXPECTED_COLUMN_ORDER)

        headers_from_sheet = all_values[0]
        data_rows = all_values[1:]

        temp_df = pd.DataFrame(data_rows, columns=headers_from_sheet)
        df = pd.DataFrame(columns=EXPECTED_COLUMN_ORDER)

        for col in EXPECTED_COLUMN_ORDER:
            if col in temp_df.columns:
                df[col] = temp_df[col]
            else:
                 if "Time" in col: df[col] = pd.NaT
                 elif col in ALL_STATUS_COLUMNS: df[col] = "No"
                 elif col == "Estimate Sent To Email" or col == "Reminder Contact Method": df[col] = "N/A"
                 else: df[col] = "N/A"

        df = df[EXPECTED_COLUMN_ORDER]

        string_cols_to_process = ['RMA', 'S/N', 'Part Number', 'SPC Code',
                                   'Description', 'Fault Comments', 'Resolution Comments',
                                   'Sender', 'Estimate Sent To Email', 'Reminder Contact Method'] + ALL_STATUS_COLUMNS
        for col in string_cols_to_process:
            if col in df.columns:
                df[col] = df[col].astype(str)
                if col in ALL_STATUS_COLUMNS:
                     df[col] = df[col].replace(['', 'nan', 'None', 'NaN', 'NONE', None, 'NaT'], 'No')
                elif col == "Estimate Sent To Email" or col == "Reminder Contact Method":
                     df[col] = df[col].replace(['', 'nan', 'None', 'NaN', 'NONE', None, 'NaT'], 'N/A')
                else:
                     df[col] = df[col].replace(['', 'nan', 'None', 'NaN', 'NONE', None, 'NaT'], 'N/A')

        for col in ALL_TIME_COLUMNS:
            if col in df.columns:
                df[col] = df[col].replace('N/A', None)
                df[col] = pd.to_datetime(df[col], errors='coerce')

        return df
    except Exception as e:
        st.error(f"An error occurred while loading data from Google Sheets: {type(e).__name__} - {e}")
    return pd.DataFrame(columns=EXPECTED_COLUMN_ORDER)


def find_row_in_gsheet(worksheet, rma_to_find, sn_to_find, headers):
    """
    Finds a row in the worksheet based on RMA and S/N.
    If RMA is missing/NA, it searches by S/N only.
    """
    try:
        rma_col_idx = headers.index("RMA")
        sn_col_idx = headers.index("S/N")
    except ValueError:
        st.error("Critical error: RMA or S/N column header not found in the Google Sheet. Cannot perform updates.")
        return -1

    all_data_values = worksheet.get_all_values()

    rma_to_find_str = str(rma_to_find).strip().lower()
    sn_to_find_str = str(sn_to_find).strip().lower()
    search_by_sn_only = rma_to_find_str in ['n/a', '']

    for i, row_values in enumerate(all_data_values[1:], start=2):
        rma_val_from_sheet = row_values[rma_col_idx] if len(row_values) > rma_col_idx else None
        sn_val_from_sheet = row_values[sn_col_idx] if len(row_values) > sn_col_idx else None

        if sn_val_from_sheet is not None:
            sheet_sn_str = str(sn_val_from_sheet).strip().lower()

            if search_by_sn_only:
                sheet_rma_str = str(rma_val_from_sheet).strip().lower()
                if sheet_sn_str == sn_to_find_str and sheet_rma_str in ['n/a', '']:
                    return i
            elif rma_val_from_sheet is not None:
                sheet_rma_str = str(rma_val_from_sheet).strip().lower()
                if sheet_rma_str == rma_to_find_str and sheet_sn_str == sn_to_find_str:
                    return i
    return -1

def update_gsheet_cells(worksheet, updates_list):
    try:
        worksheet.batch_update(updates_list)
        return True
    except Exception as e: st.error(f"An error occurred during Google Sheet batch update: {e}"); return False

def gsheet_update_wrapper(update_function, *args):
    try:
        scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/spreadsheets",
                 "https://www.googleapis.com/auth/drive.file", "https://www.googleapis.com/auth/drive"]
        creds = ServiceAccountCredentials.from_json_keyfile_name(CREDS_FILE, scope)
        client = gspread.authorize(creds)
        spreadsheet = client.open(GSHEET_NAME)
        worksheet = spreadsheet.get_worksheet(WORKSHEET_INDEX)
        headers = worksheet.row_values(1)
        if not headers: st.error("Could not read headers from main data sheet. Update failed."); return False
        return update_function(worksheet, headers, *args)
    except Exception as e: st.error(f"General error during Google Sheet operation: {type(e).__name__} - {e}"); return False

def _update_estimate_sent_in_sheet(worksheet, headers, rma, sn, sent_to_email, sent_date_obj):
    sent_time_col_name = "Estimate Sent Time"; sent_email_col_name = "Estimate Sent To Email"
    try:
        sent_time_col_idx = headers.index(sent_time_col_name) + 1
        sent_email_col_idx = headers.index(sent_email_col_name) + 1
    except ValueError: st.error(f"'{sent_time_col_name}' or '{sent_email_col_name}' not in sheet headers."); return False
    row_to_update = find_row_in_gsheet(worksheet, rma, sn, headers)
    if row_to_update != -1:
        sent_time_str = datetime.combine(sent_date_obj, datetime.now().time()).strftime("%Y-%m-%d %H:%M:%S")
        updates = [
            {'range': gspread.utils.rowcol_to_a1(row_to_update, sent_email_col_idx), 'values': [[sent_to_email]]},
            {'range': gspread.utils.rowcol_to_a1(row_to_update, sent_time_col_idx), 'values': [[sent_time_str]]} ]
        if update_gsheet_cells(worksheet, updates):
            st.success(f"Estimate for RMA {rma}, S/N {sn} marked as sent to {sent_to_email} on {sent_date_obj.strftime('%Y-%m-%d')}."); return True
    else: st.error(f"Record for RMA {rma}, S/N {sn} not found for estimate sent update.")
    return False

def _update_reminder_in_sheet(worksheet, headers, rma, sn, reminder_date_obj, contact_method):
    reminder_status_col_name = "Reminder Completed"; reminder_time_col_name = "Reminder Completed Time"
    reminder_method_col_name = "Reminder Contact Method"
    try:
        reminder_status_col_idx = headers.index(reminder_status_col_name) + 1
        reminder_time_col_idx = headers.index(reminder_time_col_name) + 1
        reminder_method_col_idx = headers.index(reminder_method_col_name) + 1
    except ValueError: st.error(f"One of the reminder columns not in sheet headers."); return False

    row_to_update = find_row_in_gsheet(worksheet, rma, sn, headers)
    if row_to_update != -1:
        reminder_time_str = datetime.combine(reminder_date_obj, datetime.now().time()).strftime("%Y-%m-%d %H:%M:%S")
        updates = [
            {'range': gspread.utils.rowcol_to_a1(row_to_update, reminder_status_col_idx), 'values': [["Yes"]]},
            {'range': gspread.utils.rowcol_to_a1(row_to_update, reminder_time_col_idx), 'values': [[reminder_time_str]]},
            {'range': gspread.utils.rowcol_to_a1(row_to_update, reminder_method_col_idx), 'values': [[contact_method]]}
        ]
        if update_gsheet_cells(worksheet, updates):
            st.success(f"Reminder for RMA {rma}, S/N {sn} (via {contact_method}) marked as completed on {reminder_date_obj.strftime('%Y-%m-%d')}."); return True
    else: st.error(f"Record for RMA {rma}, S/N {sn} not found for reminder update.")
    return False

def _update_shipped_in_sheet(worksheet, headers, rma, sn, shipped_date_obj):
    shipped_status_col_name = "Shipped"; shipped_time_col_name = "Shipped Time"
    try:
        shipped_status_col_idx = headers.index(shipped_status_col_name) + 1
        shipped_time_col_idx = headers.index(shipped_time_col_name) + 1
    except ValueError: st.error(f"'{shipped_status_col_name}' or '{shipped_time_col_name}' not in sheet headers."); return False
    row_to_update = find_row_in_gsheet(worksheet, rma, sn, headers)
    if row_to_update != -1:
        shipped_time_str = datetime.combine(shipped_date_obj, datetime.now().time()).strftime("%Y-%m-%d %H:%M:%S")
        updates = [
            {'range': gspread.utils.rowcol_to_a1(row_to_update, shipped_status_col_idx), 'values': [["Yes"]]},
            {'range': gspread.utils.rowcol_to_a1(row_to_update, shipped_time_col_idx), 'values': [[shipped_time_str]]} ]
        if update_gsheet_cells(worksheet, updates):
            st.success(f"Successfully marked RMA {rma}, S/N {sn} as shipped on {shipped_date_obj.strftime('%Y-%m-%d')}."); return True
    else: st.error(f"Record with RMA {rma} and S/N {sn} not found in Google Sheet. Update failed.")
    return False

def update_estimate_sent_details_in_gsheet(rma, sn, sent_to_email, sent_date_obj):
    return gsheet_update_wrapper(_update_estimate_sent_in_sheet, rma, sn, sent_to_email, sent_date_obj)
def update_reminder_details_in_gsheet(rma, sn, reminder_date_obj, contact_method):
    return gsheet_update_wrapper(_update_reminder_in_sheet, rma, sn, reminder_date_obj, contact_method)
def update_shipped_status_in_gsheet(rma, sn, shipped_date_obj):
    return gsheet_update_wrapper(_update_shipped_in_sheet, rma, sn, shipped_date_obj)

def display_kpis(df):
    if df.empty: return
    total_records = len(df)
    kpi_cols = { "Est. Complete": 'Estimate Complete', "Est. Approved": 'Estimate Approved',
        "Est. Sent": 'Estimate Sent To Email', "Reminders Done": 'Reminder Completed',
        "QA Approved": 'QA Approved', "Units Shipped": 'Shipped' }
    kpi_values = {"Total Records": total_records}
    for label, col_name in kpi_cols.items():
        if col_name in df.columns:
            if col_name == 'Estimate Sent To Email': kpi_values[label] = df[df[col_name].astype(str).str.lower() != 'n/a'].shape[0]
            else: kpi_values[label] = df[df[col_name].astype(str).str.lower() == 'yes'].shape[0]
        else: kpi_values[label] = 0
    cols = st.columns(len(kpi_values))
    for i, (label, value) in enumerate(kpi_values.items()): cols[i].metric(label, value)

def identify_overdue_estimates(df, days_threshold=3):
    required_cols = ['Estimate Complete Time', 'Estimate Complete', 'Estimate Sent To Email', 'RMA', 'S/N', 'SPC Code',
                     'Shipped']
    if df.empty or not all(col in df.columns for col in required_cols): return pd.DataFrame()

    df_copy = df.copy()
    df_copy['Estimate Complete Time'] = pd.to_datetime(df_copy['Estimate Complete Time'], errors='coerce')
    now = datetime.now(); overdue_items = []
    for _, row in df_copy.iterrows():
        is_estimate_complete = str(row.get('Estimate Complete', 'N/A')).lower() == 'yes'
        is_not_shipped = str(row.get('Shipped', 'N/A')).lower() in ['no', 'n/a']
        is_not_sent = str(row.get('Estimate Sent To Email', 'N/A')).lower() == 'n/a'
        complete_time = row['Estimate Complete Time']

        if is_estimate_complete and is_not_shipped and is_not_sent and pd.notna(complete_time):
            if (now - complete_time).days > days_threshold:
                rma_value = str(row.get('RMA', 'N/A'))
                bc_url = f"{BC_BASE_URL}?company={BC_COMPANY}&page={BC_PAGE_ID}&filter='{urllib.parse.quote_plus(BC_RMA_FIELD_NAME)}'%20IS%20%27{urllib.parse.quote_plus(rma_value)}%27" if rma_value not in ['N/A', ''] else None
                overdue_items.append({
                    'RMA': rma_value, 'S/N': row.get('S/N', 'N/A'), 'SPC Code': row.get('SPC Code', 'N/A'),
                    'Estimate Complete Time': row['Estimate Complete Time'].strftime('%Y-%m-%d'),
                    'Days Overdue for Sending': (now - complete_time).days,
                    BC_LINK_COL_NAME: bc_url  })
    return pd.DataFrame(overdue_items)

def identify_overdue_for_shipping(df, days_threshold=1):
    required_cols = ['QA Approved Time', 'Estimate Complete', 'Estimate Approved', 'QA Approved', 'Shipped', 'RMA', 'S/N', 'SPC Code']
    if df.empty or not all(col in df.columns for col in required_cols): return pd.DataFrame()
    df_copy = df.copy(); df_copy['QA Approved Time'] = pd.to_datetime(df_copy['QA Approved Time'], errors='coerce')
    now = datetime.now(); overdue_shipping_items = []
    for _, row in df_copy.iterrows():
        if str(row.get('Estimate Complete', 'N/A')).lower() == 'yes' and \
           str(row.get('Estimate Approved', 'N/A')).lower() == 'yes' and \
           str(row.get('QA Approved', 'N/A')).lower() == 'yes' and \
           str(row.get('Shipped', 'N/A')).lower() in ['no', 'n/a'] and \
           pd.notna(row['QA Approved Time']):
            if (now - row['QA Approved Time']).days > days_threshold:
                rma_value = str(row.get('RMA', 'N/A'))
                bc_url = f"{BC_BASE_URL}?company={BC_COMPANY}&page={BC_PAGE_ID}&filter='{urllib.parse.quote_plus(BC_RMA_FIELD_NAME)}'%20IS%20%27{urllib.parse.quote_plus(rma_value)}%27" if rma_value not in ['N/A', ''] else None
                overdue_shipping_items.append({
                    'RMA': rma_value, 'S/N': row.get('S/N', 'N/A'), 'SPC Code': row.get('SPC Code', 'N/A'),
                    'QA Approved Time': row['QA Approved Time'].strftime('%Y-%m-%d'),
                    'Days Pending Shipping': (now - row['QA Approved Time']).days, BC_LINK_COL_NAME: bc_url })
    return pd.DataFrame(overdue_shipping_items)

def identify_overdue_reminders(df, days_threshold=2):
    required_cols = ['Estimate Sent Time', 'Estimate Sent To Email', 'Reminder Completed', 'RMA', 'S/N', 'SPC Code', 'Reminder Contact Method',
                     'Estimate Approved']
    if df.empty or not all(col in df.columns for col in required_cols): return pd.DataFrame()

    df_copy = df.copy()
    df_copy['Estimate Sent Time'] = pd.to_datetime(df_copy['Estimate Sent Time'], errors='coerce')
    now = datetime.now()
    overdue_reminder_items = []

    for _, row in df_copy.iterrows():
        is_estimate_sent = str(row.get('Estimate Sent To Email', 'N/A')).lower() != 'n/a'
        is_reminder_not_done = str(row.get('Reminder Completed', 'N/A')).lower() in ['no', 'n/a']
        is_estimate_not_approved = str(row.get('Estimate Approved', 'N/A')).lower() in ['no', 'n/a']
        estimate_sent_time = row['Estimate Sent Time']; rma_value = str(row.get('RMA', 'N/A'))

        # Reminder is needed if sent, not yet approved, and reminder not done.
        if is_estimate_sent and is_reminder_not_done and is_estimate_not_approved and pd.notna(estimate_sent_time):
            days_passed_reminder = (now - estimate_sent_time).days
            if days_passed_reminder > days_threshold:
                bc_url = f"{BC_BASE_URL}?company={BC_COMPANY}&page={BC_PAGE_ID}&filter='{urllib.parse.quote_plus(BC_RMA_FIELD_NAME)}'%20IS%20%27{urllib.parse.quote_plus(rma_value)}%27" if rma_value not in ['N/A', ''] else None
                overdue_reminder_items.append({
                    'RMA': rma_value, 'S/N': row.get('S/N', 'N/A'), 'SPC Code': row.get('SPC Code', 'N/A'),
                    'Estimate Sent To Email': row.get('Estimate Sent To Email', 'N/A'),
                    'Estimate Sent Time': estimate_sent_time.strftime('%Y-%m-%d') if pd.notna(estimate_sent_time) else 'N/A',
                    'Days Pending Reminder': days_passed_reminder,
                    'Reminder Contact Method': row.get('Reminder Contact Method', 'N/A'),
                    'Estimate Approved': row.get('Estimate Approved', 'N/A'),
                    BC_LINK_COL_NAME: bc_url  })
    return pd.DataFrame(overdue_reminder_items)

# --- Daily Status Report Functions (Modified for GSheet Archive) ---
@st.cache_data(ttl=60)
def get_archived_reports_from_gsheet(archive_sheet_name, expected_headers):
    """Loads all archived reports from the specified Google Sheet archive tab."""
    try:
        scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/spreadsheets",
                 "https://www.googleapis.com/auth/drive.file", "https://www.googleapis.com/auth/drive"]
        creds = ServiceAccountCredentials.from_json_keyfile_name(CREDS_FILE, scope)
        client = gspread.authorize(creds)
        spreadsheet = client.open(GSHEET_NAME)
        try:
            archive_ws = spreadsheet.worksheet(archive_sheet_name)
        except gspread.exceptions.WorksheetNotFound:
            st.error(f"Archive sheet '{archive_sheet_name}' not found. Please create it with headers: {', '.join(expected_headers)}.")
            return []

        records = archive_ws.get_all_records()
        archived_reports = []

        if archive_sheet_name == ARCHIVE_SHEET_NAME: # Daily Status Report Archive
            for rec in records:
                try:
                    needs_est_str = rec.get('Needs Estimate Creation', '[]')
                    needs_ship_str = rec.get('Needs Shipping', '[]')
                    needs_reminder_str = rec.get('Needs Reminder', '[]')
                    needs_est_list = json.loads(needs_est_str) if needs_est_str and needs_est_str.strip() else []
                    needs_ship_list = json.loads(needs_ship_str) if needs_ship_str and needs_ship_str.strip() else []
                    needs_reminder_list = json.loads(needs_reminder_str) if needs_reminder_str and needs_reminder_str.strip() else []
                    archived_reports.append({
                        "date": rec.get('Report Date'),
                        "needs_estimate_creation": needs_est_list,
                        "needs_shipping": needs_ship_list,
                        "needs_reminder": needs_reminder_list })
                except Exception: pass # Skip malformed rows
        elif archive_sheet_name == EOD_SUMMARY_ARCHIVE_SHEET_NAME:
            for rec in records:
                try:
                    archived_reports.append({
                        "date": rec.get('Report Date'),
                        "estimate_tasks": json.loads(rec.get('Estimate Task Summary', '[]')),
                        "reminder_tasks": json.loads(rec.get('Reminder Task Summary', '[]')),
                        "shipping_tasks": json.loads(rec.get('Shipping Task Summary', '[]')),
                        "adhoc_shipped_today": json.loads(rec.get('AdHoc Shipped Today', '[]'))
                    })
                except Exception: pass # Skip malformed rows

        archived_reports.sort(key=lambda r: r.get('date', ''), reverse=True)
        return archived_reports
    except Exception as e:
        st.error(f"Error loading archived reports from '{archive_sheet_name}': {type(e).__name__} - {e}")
        return []

def get_last_report_date_from_archive(archived_reports):
    if not archived_reports: return date.today() - timedelta(days=1)
    try:
        latest_date_str = archived_reports[0]['date']
        return datetime.strptime(latest_date_str, "%Y-%m-%d").date()
    except: return date.today() - timedelta(days=1)

def save_report_to_gsheet_archive(report_data, archive_sheet_name_to_save, archive_headers_to_check):
    """Saves a single daily report to the specified Google Sheet archive."""
    try:
        scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/spreadsheets",
                 "https://www.googleapis.com/auth/drive.file", "https://www.googleapis.com/auth/drive"]
        creds = ServiceAccountCredentials.from_json_keyfile_name(CREDS_FILE, scope)
        client = gspread.authorize(creds)
        spreadsheet = client.open(GSHEET_NAME)
        try:
            archive_ws = spreadsheet.worksheet(archive_sheet_name_to_save)
            if archive_ws.row_count == 0 or (archive_ws.row_count >= 1 and archive_ws.row_values(1) != archive_headers_to_check):
                st.info(f"Resetting headers for archive sheet '{archive_sheet_name_to_save}'.")
                archive_ws.clear(); archive_ws.append_row(archive_headers_to_check)
        except gspread.exceptions.WorksheetNotFound:
            st.info(f"Archive sheet '{archive_sheet_name_to_save}' not found. Creating it with headers: {', '.join(archive_headers_to_check)}.")
            archive_ws = spreadsheet.add_worksheet(title=archive_sheet_name_to_save, rows="100", cols=str(len(archive_headers_to_check)))
            archive_ws.append_row(archive_headers_to_check)

        existing_dates = archive_ws.col_values(1)[1:]

        row_to_append = [report_data['date']]
        if archive_sheet_name_to_save == ARCHIVE_SHEET_NAME:
            if report_data['date'] in existing_dates: return False
            row_to_append.extend([
                json.dumps(report_data['needs_estimate_creation']),
                json.dumps(report_data['needs_shipping']),
                json.dumps(report_data['needs_reminder'])
            ])
        elif archive_sheet_name_to_save == EOD_SUMMARY_ARCHIVE_SHEET_NAME:
             # For EOD, if report for date exists, update it. Otherwise, append.
            row_number_to_update_eod = -1
            if report_data['date'] in existing_dates:
                try:
                    cell = archive_ws.find(report_data['date'])
                    row_number_to_update_eod = cell.row
                except gspread.exceptions.CellNotFound:
                    pass # Will append if not found

            row_to_append.extend([
                json.dumps(report_data['estimate_tasks']),
                json.dumps(report_data['reminder_tasks']),
                json.dumps(report_data['shipping_tasks']),
                json.dumps(report_data.get('adhoc_shipped_today', []))
            ])
            if row_number_to_update_eod != -1:
                # Construct list of Cell objects for batch update of the row
                cell_list = [gspread.Cell(row_number_to_update_eod, i+1, val) for i, val in enumerate(row_to_append)]
                archive_ws.update_cells(cell_list)
                st.info(f"EOD Summary for {report_data['date']} updated in archive.")
                if archive_sheet_name_to_save == ARCHIVE_SHEET_NAME: get_archived_reports_from_gsheet.clear(archive_sheet_name=ARCHIVE_SHEET_NAME, expected_headers=ARCHIVE_SHEET_HEADERS)
                elif archive_sheet_name_to_save == EOD_SUMMARY_ARCHIVE_SHEET_NAME: get_archived_reports_from_gsheet.clear(archive_sheet_name=EOD_SUMMARY_ARCHIVE_SHEET_NAME, expected_headers=EOD_ARCHIVE_SHEET_HEADERS)
                return True # Indicate update/save

        archive_ws.append_row(row_to_append)

        if archive_sheet_name_to_save == ARCHIVE_SHEET_NAME:
            get_archived_reports_from_gsheet.clear(archive_sheet_name=ARCHIVE_SHEET_NAME, expected_headers=ARCHIVE_SHEET_HEADERS)
        elif archive_sheet_name_to_save == EOD_SUMMARY_ARCHIVE_SHEET_NAME:
            get_archived_reports_from_gsheet.clear(archive_sheet_name=EOD_SUMMARY_ARCHIVE_SHEET_NAME, expected_headers=EOD_ARCHIVE_SHEET_HEADERS)
        return True
    except Exception as e: st.error(f"Error saving report to '{archive_sheet_name_to_save}': {type(e).__name__} - {e}"); return False

def generate_single_day_report_content(df, report_date_obj):
    report_content = { "date": report_date_obj.strftime("%Y-%m-%d"),
                      "needs_shipping": [],
                      "needs_estimate_creation": [],
                      "needs_reminder": [] }

    # Needs Shipping
    shipping_df = df[
        (df['Estimate Complete'].astype(str).str.lower() == 'yes') &
        (df['Estimate Approved'].astype(str).str.lower() == 'yes') &
        (df['QA Approved'].astype(str).str.lower() == 'yes') &
        (df['Shipped'].astype(str).str.lower().isin(['no', 'n/a'])) &
        (pd.to_datetime(df['QA Approved Time'], errors='coerce').dt.date == report_date_obj) ]
    for _, row in shipping_df.iterrows():
        report_content["needs_shipping"].append({'RMA': str(row['RMA']), 'S/N': str(row['S/N']), 'SPC Code': str(row.get('SPC Code', 'N/A'))})

    # Needs Estimate Creation
    day_prior_to_report = report_date_obj - timedelta(days=1)
    estimate_df = df[
        (df['Estimate Complete'].astype(str).str.lower() == 'yes') &
        (df['Estimate Sent To Email'].astype(str).str.lower() == 'n/a') &
        (pd.to_datetime(df['Estimate Complete Time'], errors='coerce').dt.date == day_prior_to_report) ]
    for _, row in estimate_df.iterrows():
        report_content["needs_estimate_creation"].append({
            'RMA': str(row['RMA']), 'S/N': str(row['S/N']), 'SPC Code': str(row.get('SPC Code', 'N/A')),
            'Est. Complete Date': day_prior_to_report.strftime('%Y-%m-%d') })

    # Needs Reminder (Estimate Sent 2 days before report_date_obj, Reminder Not Completed)
    estimate_sent_target_date = report_date_obj - timedelta(days=2)
    reminder_df = df[
        (df['Estimate Sent To Email'].astype(str).str.lower() != 'n/a') &
        (df['Reminder Completed'].astype(str).str.lower().isin(['no', 'n/a'])) &
        (df['Estimate Approved'].astype(str).str.lower().isin(['no', 'n/a'])) &
        (pd.to_datetime(df['Estimate Sent Time'], errors='coerce').dt.date == estimate_sent_target_date)
    ]
    for _, row in reminder_df.iterrows():
        report_content["needs_reminder"].append({
            'RMA': str(row['RMA']),
            'S/N': str(row['S/N']),
            'SPC Code': str(row.get('SPC Code', 'N/A')),
            'Estimate Sent To Email': str(row['Estimate Sent To Email']),
            'Estimate Sent Time': pd.to_datetime(row['Estimate Sent Time']).strftime('%Y-%m-%d') if pd.notna(row['Estimate Sent Time']) else 'N/A'
        })
    return report_content

def create_excel_report_bytes(report_data, report_type="Daily"):
    """Creates an Excel file in bytes from the structured report data with improved formatting."""
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        workbook = writer.book
        header_format = workbook.add_format({'bold': True, 'text_wrap': False, 'valign': 'top', 'fg_color': '#D7E4BC', 'border': 1, 'align': 'center'})
        cell_format = workbook.add_format({'border': 1})
        title_format = workbook.add_format({'bold': True, 'font_size': 14, 'align': 'center', 'valign': 'vcenter'})

        report_date_for_title = report_data.get('date', 'Unknown_Date')

        if report_type in ["Newly Generated", "Custom Date Report", "Archived Daily Report", "Daily", "Custom"]:
            sheets_data = {
                "Needs Estimate Creation": report_data.get("needs_estimate_creation", []),
                "Needs Reminder": report_data.get("needs_reminder", []),
                "Needs Shipping": report_data.get("needs_shipping", [])
            }
        elif report_type in ["EOD", "Archived EOD Summary"]:
            sheets_data = {
                "EOD Estimate Tasks": report_data.get("estimate_tasks", []),
                "EOD Reminder Tasks": report_data.get("reminder_tasks", []),
                "EOD Shipping Tasks": report_data.get("shipping_tasks", []),
                "EOD AdHoc Shipped": report_data.get("adhoc_shipped_today", [])
            }
        else:
            st.error(f"Unexpected report_type '{report_type}' received by create_excel_report_bytes. Cannot determine sheet structure.")
            return BytesIO().getvalue() # Return empty bytes for an error case

        for sheet_name_key, data_list in sheets_data.items():
            df_report_sheet = pd.DataFrame(data_list)

            default_cols = ['RMA', 'S/N', 'SPC Code']
            if "Estimate Creation" in sheet_name_key: default_cols.extend(['Est. Complete Date'])
            elif "Reminder" in sheet_name_key and report_type not in ["EOD", "Archived EOD Summary"]: default_cols.extend(['Estimate Sent To Email', 'Estimate Sent Time'])
            elif "Shipping" in sheet_name_key and report_type not in ["EOD", "Archived EOD Summary"]: pass
            elif ("EOD" in sheet_name_key or "EOD" in report_type) and "AdHoc" not in sheet_name_key : default_cols.extend(['Status', 'Original Task'])
            elif "AdHoc" in sheet_name_key: default_cols = ['RMA', 'S/N', 'SPC Code', 'Shipped Time']


            if not df_report_sheet.empty:
                for col in default_cols:
                    if col not in df_report_sheet.columns: df_report_sheet[col] = 'N/A'

                other_cols_present = [col for col in df_report_sheet.columns if col not in default_cols]
                final_cols_order = default_cols + other_cols_present
                final_cols_order = [col for col in final_cols_order if col in df_report_sheet.columns]
                df_report_sheet = df_report_sheet[final_cols_order]

                df_for_excel = df_report_sheet.astype(str)

                df_for_excel.to_excel(writer, sheet_name=sheet_name_key, startrow=2, index=False, header=False)
                worksheet = writer.sheets[sheet_name_key]
                worksheet.merge_range(0, 0, 0, len(df_for_excel.columns)-1 if len(df_for_excel.columns)>0 else 0, f"{sheet_name_key} - Report Date: {report_date_for_title}", title_format)
                worksheet.set_row(0, 30)
                for col_num, value in enumerate(df_for_excel.columns.values): worksheet.write(2, col_num, value, header_format)
                for row_num in range(3, len(df_for_excel) + 3):
                    for col_num in range(len(df_for_excel.columns)):
                        worksheet.write(row_num, col_num, df_for_excel.iloc[row_num-3, col_num], cell_format)
                for i, col_name_iter in enumerate(df_for_excel.columns):
                    header_len = len(str(col_name_iter))
                    data_max_len = 0
                    if not df_for_excel[col_name_iter].empty:
                        try:
                            lengths = df_for_excel[col_name_iter].map(len)
                            if not lengths.empty: data_max_len = lengths.max()
                        except : data_max_len = 0
                    column_width = max(data_max_len, header_len) + 2
                    worksheet.set_column(i, i, column_width)
            else:
                worksheet = writer.book.add_worksheet(sheet_name_key)
                worksheet.merge_range(0, 0, 0, 2, f"{sheet_name_key} - Report Date: {report_date_for_title}", title_format)
                worksheet.write(2,0, "No items for this category.", cell_format)
    return output.getvalue()

def display_formatted_report(report_data, source="Newly Generated", report_key_suffix=""):
    st.markdown(f"### {source} Daily Status Report for: {report_data['date']}")
    st.markdown(f"**📋 Needs Estimate Creation (from items completed on {(datetime.strptime(report_data['date'], '%Y-%m-%d') - timedelta(days=1)).strftime('%Y-%m-%d')}):**")
    if report_data['needs_estimate_creation']:
        for item in report_data['needs_estimate_creation']:
            st.markdown(f"- RMA: {item.get('RMA', 'N/A')}, S/N: {item.get('S/N', 'N/A')}, SPC: {item.get('SPC Code', 'N/A')} (Est. Complete: {item.get('Est. Complete Date', 'N/A')})")
    else: st.info("None for this category.")

    st.markdown(f"**📞 Needs Reminder (Estimate Sent 2 days prior to {report_data['date']}):**")
    if report_data.get('needs_reminder'):
        for item in report_data['needs_reminder']:
            st.markdown(f"- RMA: {item.get('RMA', 'N/A')}, S/N: {item.get('S/N', 'N/A')}, SPC: {item.get('SPC Code', 'N/A')}, Email: {item.get('Estimate Sent To Email', 'N/A')}, Sent Time: {item.get('Estimate Sent Time', 'N/A')}")
    else: st.info("None for this category.")

    st.markdown(f"**🚢 Needs Shipping (QA'd on {report_data['date']}):**")
    if report_data['needs_shipping']:
        for item in report_data['needs_shipping']:
            st.markdown(f"- RMA: {item.get('RMA', 'N/A')}, S/N: {item.get('S/N', 'N/A')}, SPC: {item.get('SPC Code', 'N/A')}")
    else: st.info("None for this category.")
    excel_bytes = create_excel_report_bytes(report_data, report_type=source)
    st.download_button(
        label=f"Download Report for {report_data['date']} (Excel)", data=excel_bytes,
        file_name=f"Daily_Status_Report_{report_data['date']}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=f"download_report_{report_data['date']}_{report_key_suffix}" )
    st.markdown("---")


# --- Functions from Estimate_Form_App2.py adapted for Streamlit ---

def create_estimate_files(rma, serial, contact, cust_name, cust_num, cust_desc, tech_eval, source_file_bytes):
    try:
        # Load the source workbook from bytes
        source_wb = openpyxl.load_workbook(filename=BytesIO(source_file_bytes))
        source_sheet = source_wb.active
        df = pd.read_excel(BytesIO(source_file_bytes))


        # Load the destination workbook (template)
        destination_wb = openpyxl.load_workbook('Estimate Form Template.xlsx')
        destination_sheet = destination_wb.active

        # --- Image Handling (ensure images are in a folder named 'images') ---
        image_file_path = 'images/Iridex logo.png'
        image_file_path2 = 'images/warranty.png'
        if os.path.exists(image_file_path):
            img = ExcelImage(image_file_path)
            img.width, img.height = 250, 50
            destination_sheet.add_image(img, 'D1')
        if os.path.exists(image_file_path2):
            img2 = ExcelImage(image_file_path2)
            img2.width, img2.height = 440, 22
            destination_sheet.add_image(img2, 'F6')

        # --- Data Transfer Logic ---
        column_mappings = {
            'Description': 5,
            'Quantity': 3,
            'No.': 1,
            'Amount Including Tax': 9
        }
        start_row_dest = 19
        header_row_source = 1
        source_headers = [cell.value for cell in source_sheet[header_row_source]]

        for col_name_source, col_idx_dest in column_mappings.items():
            if col_name_source in source_headers:
                col_idx_source = source_headers.index(col_name_source) + 1
                current_dest_row = start_row_dest
                for row in range(header_row_source + 1, source_sheet.max_row + 1):
                    cell_value = source_sheet.cell(row=row, column=col_idx_source).value
                    if cell_value is not None:
                         destination_sheet.cell(row=current_dest_row, column=col_idx_dest).value = cell_value
                         current_dest_row +=1


        # --- Populate other fields ---
        destination_sheet['F4'].value = rma
        destination_sheet['B4'].value = serial
        destination_sheet['G4'].value = contact
        destination_sheet['A6'].value = cust_name
        destination_sheet['C6'].value = cust_num
        destination_sheet['A9'].value = cust_desc
        destination_sheet['A11'].value = tech_eval

        # --- Calculate Total ---
        total_sum = 0
        if 'Amount Including Tax' in df.columns:
            total_sum = df['Amount Including Tax'].sum()
            
        # Save to BytesIO objects
        excel_output = BytesIO()
        destination_wb.save(excel_output)
        excel_output.seek(0)
        
        # --- Update Google Sheet ---
        update_google_sheet_estimate(rma, serial, contact, cust_name, cust_num, str(total_sum), cust_desc, tech_eval)

        return excel_output, total_sum

    except Exception as e:
        st.error(f"Error creating estimate files: {e}")
        return None, 0


def next_row_gsheet(sh):
    values = sh.col_values(1)  # Get all values in the first column
    return len(values) + 1  # Return the next row number

def update_google_sheet_estimate(rma, serial, contact, cust_name, cust_num, est_val, cust_desc, tech_eval):
    try:
        scope = ["https://spreadsheets.google.com/feeds",'https://www.googleapis.com/auth/spreadsheets',"https://www.googleapis.com/auth/drive","https://www.googleapis.com/auth/drive.file"]
        creds = ServiceAccountCredentials.from_json_keyfile_name("Credentials.json", scope)
        client = gspread.authorize(creds)
        sh = client.open("Estimate form").get_worksheet(0)

        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        cell = sh.find(rma)

        if cell is None:
            next_cell = next_row_gsheet(sh)
            sh.update_acell(f"A{next_cell}", rma)
            sh.update_acell(f"B{next_cell}", serial)
            sh.update_acell(f"C{next_cell}", contact)
            sh.update_acell(f"D{next_cell}", cust_name)
            sh.update_acell(f"E{next_cell}", cust_num)
            sh.update_acell(f"F{next_cell}", timestamp)
            sh.update_acell(f"G{next_cell}", est_val)
            sh.update_acell(f"H{next_cell}", cust_desc)
            sh.update_acell(f"I{next_cell}", tech_eval)
            sh.update_acell(f"K{next_cell}", timestamp)
        else:
            row = cell.row
            sh.update_acell(f"A{row}", rma)
            sh.update_acell(f"B{row}", serial)
            sh.update_acell(f"C{row}", contact)
            sh.update_acell(f"D{row}", cust_name)
            sh.update_acell(f"E{row}", cust_num)
            sh.update_acell(f"F{row}", timestamp)
            sh.update_acell(f"G{row}", est_val)
            sh.update_acell(f"H{row}", cust_desc)
            sh.update_acell(f"I{row}", tech_eval)
            sh.update_acell(f"K{row}", timestamp)
        st.success("Google Sheet updated successfully!")
    except Exception as e:
        st.error(f"Failed to update Google Sheet: {e}")


def insert_rma_received(rma, serial, items_list):
    try:
        scope = ["https://spreadsheets.google.com/feeds",'https://www.googleapis.com/auth/spreadsheets',"https://www.googleapis.com/auth/drive","https://www.googleapis.com/auth/drive.file"]
        creds = ServiceAccountCredentials.from_json_keyfile_name("Credentials.json", scope)
        client = gspread.authorize(creds)
        sh = client.open("Estimate form").get_worksheet(0)

        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        items_str = ", ".join(items_list)

        cell = sh.find(rma)
        if cell is None:
            next_cell = next_row_gsheet(sh)
            sh.update_acell(f"A{next_cell}", rma)
            sh.update_acell(f"B{next_cell}", serial)
            sh.update_acell(f"F{next_cell}", timestamp)
            sh.update_acell(f"J{next_cell}", timestamp)
            sh.update_acell(f"V{next_cell}", items_str)
        else:
            row = cell.row
            sh.update_acell(f"A{row}", rma)
            sh.update_acell(f"B{row}", serial)
            sh.update_acell(f"F{row}", timestamp)
            sh.update_acell(f"J{row}", timestamp)
            sh.update_acell(f"V{row}", items_str)
        st.success(f"RMA {rma} successfully inserted/updated.")
    except Exception as e:
        st.error(f"Failed to insert RMA: {e}")


def insert_loaner_details(rma, loaner_serial, tracking_number):
    try:
        scope = ["https://spreadsheets.google.com/feeds",'https://www.googleapis.com/auth/spreadsheets',"https://www.googleapis.com/auth/drive","https://www.googleapis.com/auth/drive.file"]
        creds = ServiceAccountCredentials.from_json_keyfile_name("Credentials.json", scope)
        client = gspread.authorize(creds)
        sh = client.open("Estimate form").get_worksheet(0)
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        cell = sh.find(rma)
        if cell is not None:
            row = cell.row
            sh.update_acell(f"N{row}", loaner_serial)
            sh.update_acell(f"O{row}", timestamp)
            sh.update_acell(f"P{row}", tracking_number)
            st.success(f"Loaner details for RMA {rma} updated.")
        else:
            st.warning(f"RMA {rma} not found to update loaner details.")
    except Exception as e:
        st.error(f"Failed to insert loaner details: {e}")


# --- Main Application ---
st.title("🛠️ Service & Estimate Dashboard")
st.markdown("Monitor statuses, create estimates, and manage service items.")


# Initialize session state variables
if 'first_load_complete' not in st.session_state: st.session_state.first_load_complete = False
if 'refresh_counter' not in st.session_state: st.session_state.refresh_counter = 0
if 'data_df' not in st.session_state:
    st.session_state.data_df = load_data_from_google_sheet()
    st.session_state.first_load_complete = True
if 'newly_generated_reports_to_display' not in st.session_state: st.session_state.newly_generated_reports_to_display = []
if 'selected_archived_report_to_display' not in st.session_state: st.session_state.selected_archived_report_to_display = None
if 'custom_report_to_display' not in st.session_state: st.session_state.custom_report_to_display = None
if 'end_of_day_summary_report' not in st.session_state: st.session_state.end_of_day_summary_report = None
if 'selected_eod_summary_to_display' not in st.session_state: st.session_state.selected_eod_summary_to_display = None
if 'estimate_excel_bytes' not in st.session_state: st.session_state.estimate_excel_bytes = None
if 'estimate_rma' not in st.session_state: st.session_state.estimate_rma = ""


if st.button("🔄 Refresh Data from Google Sheet"):
    load_data_from_google_sheet.clear()
    st.session_state.data_df = load_data_from_google_sheet()
    get_archived_reports_from_gsheet.clear(archive_sheet_name=ARCHIVE_SHEET_NAME, expected_headers=ARCHIVE_SHEET_HEADERS)
    get_archived_reports_from_gsheet.clear(archive_sheet_name=EOD_SUMMARY_ARCHIVE_SHEET_NAME, expected_headers=EOD_ARCHIVE_SHEET_HEADERS)
    st.session_state.first_load_complete = False
    st.session_state.refresh_counter += 1
    st.session_state.newly_generated_reports_to_display = []
    st.session_state.selected_archived_report_to_display = None
    st.session_state.custom_report_to_display = None
    st.session_state.end_of_day_summary_report = None
    st.session_state.selected_eod_summary_to_display = None
    st.rerun()

data_df = st.session_state.data_df
archived_daily_reports_gsheet = get_archived_reports_from_gsheet(ARCHIVE_SHEET_NAME, ARCHIVE_SHEET_HEADERS)
archived_eod_summaries_gsheet = get_archived_reports_from_gsheet(EOD_SUMMARY_ARCHIVE_SHEET_NAME, EOD_ARCHIVE_SHEET_HEADERS)


# --- Sidebar ---
st.sidebar.markdown("---")
st.sidebar.header("📝 Create Estimate")

with st.sidebar.expander("Estimate Form", expanded=False):
    with st.form("estimate_form", clear_on_submit=True):
        rma_est = st.text_input("RMA No.")
        serial_est = st.text_input("Serial No.")
        contact_est = st.text_input("Service Contact")
        cust_name_est = st.text_input("Customer Name")
        cust_num_est = st.text_input("Customer Number")
        cust_desc_est = st.text_area("Customer Description of Problem")
        tech_eval_est = st.text_area("Technician Product Evaluation")
        source_file_est = st.file_uploader("Upload Source Excel File", type=["xlsx"])
        submitted_est = st.form_submit_button("Create Estimate")

        if submitted_est and all([rma_est, serial_est, source_file_est]):
            source_file_bytes = source_file_est.getvalue()
            excel_bytes, total_sum = create_estimate_files(
                rma_est, serial_est, contact_est, cust_name_est, cust_num_est,
                cust_desc_est, tech_eval_est, source_file_bytes
            )
            if excel_bytes:
                 st.session_state.estimate_excel_bytes = excel_bytes
                 st.session_state.estimate_rma = rma_est
                 st.success(f"Estimate created for RMA {rma_est} with a total of ${total_sum:,.2f}. You can now download the file.")
                 st.rerun()

if st.session_state.estimate_excel_bytes:
    st.sidebar.download_button(
        label=f"Download Estimate for RMA {st.session_state.estimate_rma}",
        data=st.session_state.estimate_excel_bytes,
        file_name=f"Estimate_{st.session_state.estimate_rma}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    if st.sidebar.button("Clear Download Link"):
        st.session_state.estimate_excel_bytes = None
        st.session_state.estimate_rma = ""
        st.rerun()


st.sidebar.markdown("---")
st.sidebar.header("📬 Received RMA")
with st.sidebar.expander("Insert Received RMA", expanded=False):
    with st.form("received_rma_form", clear_on_submit=True):
        rma_received = st.text_input("RMA No.", key="rma_rec")
        serial_received = st.text_input("Serial No.", key="serial_rec")
        st.write("Items Received:")
        footswitch = st.checkbox("Footswitch")
        remote_interlock = st.checkbox("Remote Interlock")
        keys = st.checkbox("Keys")
        case = st.checkbox("Case")
        smartkey = st.checkbox("Smart Key")
        misc_text = st.text_input("Misc Item (if any)")

        submitted_received = st.form_submit_button("Insert RMA")
        if submitted_received and rma_received and serial_received:
            items = []
            if footswitch: items.append("Footswitch")
            if remote_interlock: items.append("Remote Interlock")
            if keys: items.append("Keys")
            if case: items.append("Case")
            if smartkey: items.append("Smart Key")
            if misc_text: items.append(misc_text)
            insert_rma_received(rma_received, serial_received, items)

st.sidebar.markdown("---")
st.sidebar.header("🚚 Insert Loaner")
with st.sidebar.expander("Insert Loaner Details", expanded=False):
     with st.form("loaner_form", clear_on_submit=True):
        rma_loaner = st.text_input("RMA No.", key="rma_loan")
        loaner_serial = st.text_input("Loaner Serial No.", key="loaner_serial")
        tracking_loaner = st.text_input("Tracking Number", key="tracking_loaner")
        submitted_loaner = st.form_submit_button("Insert Loaner")

        if submitted_loaner and rma_loaner and loaner_serial:
            insert_loaner_details(rma_loaner, loaner_serial, tracking_loaner)


# The rest of your dashboard code continues here...
if not data_df.empty:
    st.sidebar.markdown("---")
    st.sidebar.header("📅 Daily Status Reports")
    if st.sidebar.button("Generate Daily Status Report(s)", key=f"gen_daily_report_btn_{st.session_state.refresh_counter}"):
        st.session_state.newly_generated_reports_to_display = []
        st.session_state.selected_archived_report_to_display = None
        st.session_state.custom_report_to_display = None
        st.session_state.end_of_day_summary_report = None
        st.session_state.selected_eod_summary_to_display = None
        if data_df.empty:
            st.sidebar.warning("No data loaded to generate reports.")
        else:
            last_gen_date_from_archive = get_last_report_date_from_archive(archived_daily_reports_gsheet)
            today = date.today()
            current_date_to_report = last_gen_date_from_archive + timedelta(days=1)
            reports_generated_this_run = []
            if current_date_to_report > today:
                 st.sidebar.info("Daily reports are up to date according to archive.")
            else:
                while current_date_to_report <= today:
                    report_data = generate_single_day_report_content(data_df, current_date_to_report)
                    if save_report_to_gsheet_archive(report_data, ARCHIVE_SHEET_NAME, ARCHIVE_SHEET_HEADERS):
                        reports_generated_this_run.append(report_data)
                        get_archived_reports_from_gsheet.clear(archive_sheet_name=ARCHIVE_SHEET_NAME, expected_headers=ARCHIVE_SHEET_HEADERS)
                    if current_date_to_report == today: break
                    current_date_to_report += timedelta(days=1)
                    if (current_date_to_report - (last_gen_date_from_archive + timedelta(days=1))).days > 30 :
                        st.sidebar.error("More than 30 days of reports to generate."); break
                if reports_generated_this_run:
                    st.session_state.newly_generated_reports_to_display = reports_generated_this_run
                    st.sidebar.success(f"{len(reports_generated_this_run)} daily report(s) generated and saved to Google Sheet archive.")
                elif last_gen_date_from_archive >= today :
                    st.sidebar.info("Daily report for today already in archive or no new days to report.")
                st.rerun()

    # (The rest of the original dashboard code for reports, logging, etc. follows)
    # ... (omitted for brevity but should be included in the final script)
else:
    st.info("No data to display. Please ensure the Google Sheet is accessible, contains data with headers for all expected columns, and 'Credentials.json' is correctly set up.")

st.markdown("---")
st.markdown("Built with ❤️ using [Streamlit](https://streamlit.io) and Google Sheets")
# --- MAIN DISPLAY AREA ---
# --- Display Sections ---
if st.session_state.newly_generated_reports_to_display:
    st.markdown("---"); st.subheader("✨ Newly Generated Daily Status Report(s)")
    for i, report in enumerate(st.session_state.newly_generated_reports_to_display):
        display_formatted_report(report, source="Newly Generated", report_key_suffix=f"new_{i}")
    if st.button("Clear Newly Generated Reports View", key="clear_new_reports"):
        st.session_state.newly_generated_reports_to_display = []; st.rerun()
    st.markdown("---")

if st.session_state.selected_archived_report_to_display: # For Daily Status Archive
    st.markdown("---")
    display_formatted_report(st.session_state.selected_archived_report_to_display, source="Archived Daily Report", report_key_suffix="archive_daily_disp")
    if st.button("Close Archived Daily Report View", key="close_archive_daily_view"):
        st.session_state.selected_archived_report_to_display = None; st.rerun()
    st.markdown("---")

st.markdown("---")
st.subheader("🔍 Generate Custom Date Status Report")
custom_report_date_val = st.date_input("Select Date for Custom Report:", value=date.today(), key=f"custom_report_date_picker_{st.session_state.refresh_counter}")
if st.button("Generate Report for Selected Date", key=f"gen_custom_report_btn_{st.session_state.refresh_counter}"):
    if data_df.empty: st.warning("No data loaded to generate a custom report.")
    elif custom_report_date_val:
        st.session_state.custom_report_to_display = generate_single_day_report_content(data_df, custom_report_date_val)
        st.session_state.newly_generated_reports_to_display = []; st.session_state.selected_archived_report_to_display = None; st.session_state.end_of_day_summary_report = None; st.session_state.selected_eod_summary_to_display = None; st.rerun()
    else: st.warning("Please select a date for the custom report.")

if st.session_state.custom_report_to_display:
    st.markdown("---")
    display_formatted_report(st.session_state.custom_report_to_display, source="Custom Date Report", report_key_suffix="custom_disp")
    if st.button("Clear Custom Report View", key="clear_custom_report"):
        st.session_state.custom_report_to_display = None; st.rerun()
    st.markdown("---")

if st.session_state.get('end_of_day_summary_report'): # For Live EOD Summary
    eod_summary = st.session_state.end_of_day_summary_report
    st.markdown("---"); st.subheader(f"🏁 End of Day Summary for: {eod_summary['date']}")
    eod_display_cols = ["RMA", "S/N", "SPC Code", "Status", "Original Task"]; adhoc_shipped_cols = ["RMA", "S/N", "SPC Code", "Shipped Time"]
    st.markdown("**Estimate Creation Task Summary:**")
    if eod_summary['estimate_tasks']:
        eod_est_df = pd.DataFrame(eod_summary['estimate_tasks'])
        for col in eod_display_cols:
            if col not in eod_est_df.columns: eod_est_df[col] = "N/A"
        st.dataframe(eod_est_df[eod_display_cols], use_container_width=True)
    else: st.info("No estimate creation tasks were on today's daily report or daily report not generated for today.")
    st.markdown("**Reminder Task Summary:**")
    if eod_summary.get('reminder_tasks'):
        eod_rem_df = pd.DataFrame(eod_summary['reminder_tasks'])
        for col in eod_display_cols:
            if col not in eod_rem_df.columns: eod_rem_df[col] = "N/A"
        st.dataframe(eod_rem_df[eod_display_cols], use_container_width=True)
    else: st.info("No reminder tasks were on today's daily report or daily report not generated for today.")
    st.markdown("**Shipping Task Summary (from Daily Report):**")
    if eod_summary['shipping_tasks']:
        eod_ship_df = pd.DataFrame(eod_summary['shipping_tasks'])
        for col in eod_display_cols:
            if col not in eod_ship_df.columns: eod_ship_df[col] = "N/A"
        st.dataframe(eod_ship_df[eod_display_cols], use_container_width=True)
    else: st.info("No shipping tasks were on today's daily report or daily report not generated for today.")
    st.markdown("**Ad-hoc Shipped Today (not on initial daily report):**")
    if eod_summary.get('adhoc_shipped_today'):
        eod_adhoc_df = pd.DataFrame(eod_summary['adhoc_shipped_today'])
        for col in adhoc_shipped_cols:
            if col not in eod_adhoc_df.columns: eod_adhoc_df[col] = "N/A"
        st.dataframe(eod_adhoc_df[adhoc_shipped_cols], use_container_width=True)
    else: st.info("No additional items were marked as shipped today outside of the daily report tasks.")
    eod_output = BytesIO()
    with pd.ExcelWriter(eod_output, engine='xlsxwriter') as writer:
        workbook = writer.book
        header_format = workbook.add_format({'bold': True, 'text_wrap': False, 'valign': 'top', 'fg_color': '#D7E4BC', 'border': 1, 'align': 'center'})
        cell_format = workbook.add_format({'border': 1})
        title_format = workbook.add_format({'bold': True, 'font_size': 14, 'align': 'center', 'valign': 'vcenter'})
        eod_sheets_data = {
            "EOD Estimate Tasks": eod_summary.get("estimate_tasks", []),
            "EOD Reminder Tasks": eod_summary.get("reminder_tasks", []),
            "EOD Shipping Tasks": eod_summary.get("shipping_tasks", []),
            "EOD AdHoc Shipped": eod_summary.get("adhoc_shipped_today", [])
        }
        for sheet_name_key, data_list in eod_sheets_data.items():
            df_eod_sheet = pd.DataFrame(data_list)
            current_display_cols = eod_display_cols if "AdHoc" not in sheet_name_key else adhoc_shipped_cols
            if not df_eod_sheet.empty:
                for col in current_display_cols:
                    if col not in df_eod_sheet.columns: df_eod_sheet[col] = "N/A"
                df_eod_sheet = df_eod_sheet[current_display_cols]
            if not df_eod_sheet.empty:
                df_eod_sheet.to_excel(writer, sheet_name=sheet_name_key, startrow=2, index=False, header=False)
                ws = writer.sheets[sheet_name_key]
                ws.merge_range(0,0,0, len(df_eod_sheet.columns)-1 if len(df_eod_sheet.columns)>0 else 0, f"{sheet_name_key} - {eod_summary['date']}", title_format)
                ws.set_row(0,30)
                for cn, val in enumerate(df_eod_sheet.columns.values): ws.write(2, cn, val, header_format)
                for rn in range(3, len(df_eod_sheet)+3):
                    for cn_idx in range(len(df_eod_sheet.columns)): ws.write(rn, cn_idx, df_eod_sheet.iloc[rn-3, cn_idx], cell_format)
                for i, col in enumerate(df_eod_sheet.columns):
                    col_len = max(df_eod_sheet[col].astype(str).map(len).max(), len(col)) + 2 if not df_eod_sheet[col].empty else len(col)+2
                    ws.set_column(i,i,col_len)
            else:
                ws = writer.book.add_worksheet(sheet_name_key)
                ws.merge_range(0,0,0,2, f"{sheet_name_key} - {eod_summary['date']}", title_format)
                ws.write(2,0, f"No tasks for this category on {eod_summary['date']}.", cell_format)
    st.download_button(label=f"Download End of Day Summary ({eod_summary['date']})", data=eod_output.getvalue(),
                       file_name=f"EndOfDay_Summary_{eod_summary['date']}.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       key=f"download_eod_summary_{eod_summary['date']}_{st.session_state.refresh_counter}")
    if st.button("Clear End of Day Summary View", key=f"clear_eod_summary_{st.session_state.refresh_counter}"):
        st.session_state.end_of_day_summary_report = None; st.rerun()
    st.markdown("---")

# Display Selected Archived EOD Summary
if st.session_state.get('selected_eod_summary_to_display'):
    eod_summary_to_show = st.session_state.selected_eod_summary_to_display
    st.markdown("---")
    st.subheader(f"Archived End of Day Summary for: {eod_summary_to_show.get('date', 'N/A')}")

    eod_display_cols = ["RMA", "S/N", "SPC Code", "Status", "Original Task"]
    adhoc_shipped_cols = ["RMA", "S/N", "SPC Code", "Shipped Time"]

    st.markdown("**Estimate Creation Task Summary:**")
    if eod_summary_to_show.get('estimate_tasks'):
        eod_est_df = pd.DataFrame(eod_summary_to_show['estimate_tasks'])
        for col in eod_display_cols:
            if col not in eod_est_df.columns: eod_est_df[col] = "N/A"
        st.dataframe(eod_est_df[eod_display_cols], use_container_width=True)
    else: st.info("No estimate creation tasks in this archived summary.")

    st.markdown("**Reminder Task Summary:**")
    if eod_summary_to_show.get('reminder_tasks'):
        eod_rem_df = pd.DataFrame(eod_summary_to_show['reminder_tasks'])
        for col in eod_display_cols:
            if col not in eod_rem_df.columns: eod_rem_df[col] = "N/A"
        st.dataframe(eod_rem_df[eod_display_cols], use_container_width=True)
    else: st.info("No reminder tasks in this archived summary.")

    st.markdown("**Shipping Task Summary (from Daily Report):**")
    if eod_summary_to_show.get('shipping_tasks'):
        eod_ship_df = pd.DataFrame(eod_summary_to_show['shipping_tasks'])
        for col in eod_display_cols:
            if col not in eod_ship_df.columns: eod_ship_df[col] = "N/A"
        st.dataframe(eod_ship_df[eod_display_cols], use_container_width=True)
    else: st.info("No shipping tasks from daily report in this archived summary.")

    st.markdown("**Ad-hoc Shipped Today (not on initial daily report):**")
    if eod_summary_to_show.get('adhoc_shipped_today'):
        eod_adhoc_df = pd.DataFrame(eod_summary_to_show['adhoc_shipped_today'])
        for col in adhoc_shipped_cols:
            if col not in eod_adhoc_df.columns: eod_adhoc_df[col] = "N/A"
        st.dataframe(eod_adhoc_df[adhoc_shipped_cols], use_container_width=True)
    else: st.info("No additional items were marked as shipped ad-hoc in this archived summary.")

    excel_bytes_eod_archive = create_excel_report_bytes(eod_summary_to_show, report_type="EOD") # Use "EOD" type
    st.download_button(
        label=f"Download EOD Summary for {eod_summary_to_show.get('date', 'N/A')} (Excel)",
        data=excel_bytes_eod_archive,
        file_name=f"EOD_Summary_{eod_summary_to_show.get('date', 'N/A')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=f"download_archived_eod_{eod_summary_to_show.get('date', 'N/A')}_{st.session_state.refresh_counter}"
    )
    if st.button("Close Archived EOD Summary View", key=f"close_archived_eod_view_{st.session_state.refresh_counter}"):
        st.session_state.selected_eod_summary_to_display = None
        st.rerun()
    st.markdown("---")


if not data_df.empty:
    st.subheader("📊 Key Metrics")
    display_kpis(data_df.copy()); st.markdown("---")
    st.subheader("⚠️ Overdue Estimates Report (Pending Sending > 3 Days)")
    overdue_estimates_df = identify_overdue_estimates(data_df, days_threshold=3)
    if not overdue_estimates_df.empty:
        st.warning("The following estimates were completed more than 3 days ago and have not been sent:")
        overdue_estimates_display_cols = ['RMA', 'S/N', 'SPC Code', 'Estimate Complete Time', 'Days Overdue for Sending', BC_LINK_COL_NAME]
        st.dataframe(overdue_estimates_df[overdue_estimates_display_cols], use_container_width=True,
            column_config={ BC_LINK_COL_NAME: st.column_config.LinkColumn(label="Business Central", display_text="Open RMA")},
            column_order=overdue_estimates_display_cols )
    else: st.success("✅ No estimates are currently overdue for sending.")
    st.markdown("---")

    st.subheader("🗣️ Overdue Reminders Report (Estimate Sent > 2 Days, Awaiting Approval)")
    overdue_reminders_df = identify_overdue_reminders(data_df, days_threshold=2)
    if not overdue_reminders_df.empty:
        st.info("The following items had estimates sent >2 days ago and are pending a reminder (and are not yet approved):") # Updated title
        overdue_reminders_display_cols = ['RMA', 'S/N', 'SPC Code', 'Estimate Sent To Email', 'Estimate Sent Time', 'Days Pending Reminder', 'Reminder Contact Method', 'Estimate Approved', BC_LINK_COL_NAME]
        for col in overdue_reminders_display_cols:
            if col not in overdue_reminders_df.columns:
                overdue_reminders_df[col] = None
        st.dataframe(
            overdue_reminders_df[overdue_reminders_display_cols],
            use_container_width=True,
            column_config={
                BC_LINK_COL_NAME: st.column_config.LinkColumn(label="Business Central", display_text="Open RMA")
            },
            column_order=overdue_reminders_display_cols
        )
    else:
        st.success("✅ No items are currently overdue for a reminder.")
    st.markdown("---")


    st.subheader("🚚 Overdue for Shipping Report (QA Approved > 1 Day, Not Shipped)")
    overdue_shipping_df = identify_overdue_for_shipping(data_df, days_threshold=1)
    if not overdue_shipping_df.empty:
        st.error("The following items are QA Approved for more than 1 day and are pending shipment:")
        overdue_shipping_display_cols = ['RMA', 'S/N', 'SPC Code', 'QA Approved Time', 'Days Pending Shipping', BC_LINK_COL_NAME]
        if BC_LINK_COL_NAME not in overdue_shipping_df.columns: overdue_shipping_df[BC_LINK_COL_NAME] = None
        st.dataframe(overdue_shipping_df[overdue_shipping_display_cols], use_container_width=True,
            column_config={BC_LINK_COL_NAME: st.column_config.LinkColumn(label="Business Central", display_text="Open RMA")},
            column_order=overdue_shipping_display_cols)
    else: st.success("✅ No items are currently overdue for shipping beyond 1 day.")
    st.markdown("---")

    st.subheader("🗂️ Daily Status Report Archive")
    if archived_daily_reports_gsheet:
        report_dates = sorted(list(set(r['date'] for r in archived_daily_reports_gsheet)), reverse=True)
        available_months = sorted(list(set(datetime.strptime(d, "%Y-%m-%d").strftime("%Y-%m") for d in report_dates)), reverse=True)
        if available_months:
            selected_month_archive = st.selectbox("View Daily Reports for Month:", ["All"] + available_months, key="archive_daily_month_select")
            reports_to_list = [r for r in archived_daily_reports_gsheet if selected_month_archive == "All" or datetime.strptime(r['date'], "%Y-%m-%d").strftime("%Y-%m") == selected_month_archive]
            if reports_to_list:
                for i, report_data_item in enumerate(reports_to_list):
                    col1, col2 = st.columns([3,1])
                    with col1: st.markdown(f"**Daily Report for: {report_data_item['date']}**")
                    with col2:
                        if st.button("View/Download Daily Report", key=f"view_archive_daily_{report_data_item['date']}_{i}"):
                            st.session_state.selected_archived_report_to_display = report_data_item
                            st.session_state.newly_generated_reports_to_display = []; st.session_state.custom_report_to_display = None; st.session_state.end_of_day_summary_report = None; st.session_state.selected_eod_summary_to_display = None; st.rerun()
            else: st.info(f"No daily reports found for {selected_month_archive} in the Google Sheet archive.")
        else: st.info("No archived daily reports available in the Google Sheet yet.")
    else: st.info("No archived daily reports available in the Google Sheet yet. Generate reports using the button in the sidebar.")
    st.markdown("---")

    st.subheader("🗂️ End of Day Summary Archive")
    if archived_eod_summaries_gsheet:
        eod_report_dates = sorted(list(set(r['date'] for r in archived_eod_summaries_gsheet)), reverse=True)
        eod_available_months = sorted(list(set(datetime.strptime(d, "%Y-%m-%d").strftime("%Y-%m") for d in eod_report_dates)), reverse=True)
        if eod_available_months:
            selected_eod_month_archive = st.selectbox("View EOD Summaries for Month:", ["All"] + eod_available_months, key="archive_eod_month_select")
            eod_summaries_to_list = [r for r in archived_eod_summaries_gsheet if selected_eod_month_archive == "All" or datetime.strptime(r['date'], "%Y-%m-%d").strftime("%Y-%m") == selected_eod_month_archive]
            if eod_summaries_to_list:
                for i, eod_summary_item in enumerate(eod_summaries_to_list):
                    col1, col2 = st.columns([3,1])
                    with col1: st.markdown(f"**EOD Summary for: {eod_summary_item['date']}**")
                    with col2:
                        if st.button("View/Download EOD Summary", key=f"view_archive_eod_{eod_summary_item['date']}_{i}"):
                            st.session_state.selected_eod_summary_to_display = eod_summary_item
                            st.session_state.newly_generated_reports_to_display = []
                            st.session_state.selected_archived_report_to_display = None
                            st.session_state.custom_report_to_display = None
                            st.session_state.end_of_day_summary_report = None
                            st.rerun()
            else: st.info(f"No EOD summaries found for {selected_eod_month_archive} in the Google Sheet archive.")
        else: st.info("No archived EOD summaries available in the Google Sheet yet.")
    else: st.info("No archived EOD summaries available in the Google Sheet yet.")
    st.markdown("---")


    st.sidebar.header("🔍 Filter Options")
    filtered_df = data_df.copy()
    for col_name, search_label in [('RMA', "RMA"), ('S/N', "S/N"), ('Part Number', "Part Number"), ('SPC Code', "SPC Code")]:
        if col_name in filtered_df.columns:
            search_term = st.sidebar.text_input(f"Search by {search_label}", key=f"search_{col_name}_{st.session_state.refresh_counter}")
            if search_term: filtered_df = filtered_df[filtered_df[col_name].astype(str).str.contains(search_term, case=False, na=False)]
    status_columns_to_filter = {
        'Estimate Complete': 'Estimate Complete', 'Estimate Approved': 'Estimate Approved',
        'Reminder Completed': 'Reminder Completed', 'QA Approved': 'QA Approved', 'Shipped': 'Shipped' }
    for display_name, col_name in status_columns_to_filter.items():
        if col_name in data_df.columns:
            unique_values = ['All'] + sorted(list(set(val for val in data_df[col_name].astype(str).unique() if val and val.strip() != '' and val != 'N/A')))
            if 'N/A' in data_df[col_name].astype(str).unique(): unique_values.insert(1, "N/A")
            if 'No' in data_df[col_name].astype(str).unique() and 'No' not in unique_values : unique_values.insert(1, "No")
            default_index = 0
            if not st.session_state.first_load_complete:
                if col_name == "Shipped" and "N/A" in unique_values: default_index = unique_values.index("N/A")
                elif col_name == "Reminder Completed" and "No" in unique_values: default_index = unique_values.index("No")
            current_key = f"select_{col_name}_{st.session_state.refresh_counter}"
            selected_status = st.sidebar.selectbox(f"Filter by {display_name}", unique_values, key=current_key, index=default_index)
            if selected_status != "All":
                if col_name in filtered_df.columns: filtered_df = filtered_df[filtered_df[col_name].astype(str) == selected_status]
    st.sidebar.markdown("---"); st.sidebar.subheader("Date Range Filters")
    date_filter_columns_to_filter = {
        'Estimate Complete Time': 'Estimate Complete Time', 'Estimate Approved Time': 'Estimate Approved Time',
        'Estimate Sent Time': 'Estimate Sent Time', 'Reminder Completed Time': 'Reminder Completed Time',
        'QA Approved Time': 'QA Approved Time', 'Shipped Time': 'Shipped Time' }
    for display_name, col_name in date_filter_columns_to_filter.items():
        min_val_for_widget_setup = None; max_val_for_widget_setup = None; can_setup_widget = False
        if col_name in data_df.columns and pd.api.types.is_datetime64_any_dtype(data_df[col_name]):
            original_col_for_widget_params = data_df[col_name].dropna()
            if not original_col_for_widget_params.empty:
                min_val_for_widget_setup = original_col_for_widget_params.min().date(); max_val_for_widget_setup = original_col_for_widget_params.max().date()
                can_setup_widget = True
        if can_setup_widget:
            current_key_date = f"date_range_{col_name}_{st.session_state.refresh_counter}"
            current_date_range_selection = st.sidebar.date_input(f"Filter by {display_name}", value=[],
                min_value=min_val_for_widget_setup, max_value=max_val_for_widget_setup, key=current_key_date)
            if current_date_range_selection and len(current_date_range_selection) == 2:
                if col_name in filtered_df.columns and pd.api.types.is_datetime64_any_dtype(filtered_df[col_name]):
                    start_date_selected, end_date_selected = current_date_range_selection
                    start_datetime_selected = pd.to_datetime(start_date_selected); end_datetime_selected = pd.to_datetime(end_date_selected).replace(hour=23, minute=59, second=59)
                    condition = ((filtered_df[col_name] >= start_datetime_selected) & (filtered_df[col_name] <= end_datetime_selected) & (filtered_df[col_name].notna()) )
                    filtered_df = filtered_df[condition]
    if not st.session_state.first_load_complete: st.session_state.first_load_complete = True



    st.subheader("Filtered Data View")
    st.markdown(f"Displaying **{len(filtered_df)}** records out of **{len(data_df) if not data_df.empty else 0}** total records.")
    if not filtered_df.empty:
        df_for_display = filtered_df.copy()
        df_for_display[BC_LINK_COL_NAME] = df_for_display.apply(
            lambda row: f"{BC_BASE_URL}?company={BC_COMPANY}&page={BC_PAGE_ID}&filter='{urllib.parse.quote_plus(BC_RMA_FIELD_NAME)}'%20IS%20%27{urllib.parse.quote_plus(str(row['RMA']))}%27"
            if pd.notna(row['RMA']) and str(row['RMA']).strip() != 'N/A' and str(row['RMA']).strip() != "" else None, axis=1 )
        display_cols_order = EXPECTED_COLUMN_ORDER[:]
        if 'RMA' in display_cols_order: display_cols_order.insert(display_cols_order.index('RMA') + 1, BC_LINK_COL_NAME)
        else: display_cols_order.append(BC_LINK_COL_NAME)
        final_display_columns = [col for col in display_cols_order if col in df_for_display.columns]
        st.dataframe(df_for_display[final_display_columns], use_container_width=True,
            column_config={ BC_LINK_COL_NAME: st.column_config.LinkColumn(label="Business Central", display_text="Open RMA")})
    else: st.warning("No data matches the current filter criteria or no data loaded.")

    if not filtered_df.empty:
        st.sidebar.markdown("---"); st.sidebar.subheader("Download Data")
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            display_cols_download = [col for col in EXPECTED_COLUMN_ORDER if col in filtered_df.columns]
            df_to_export = filtered_df[display_cols_download].copy()
            for col in df_to_export.columns:
                if pd.api.types.is_datetime64_any_dtype(df_to_export[col]):
                    if df_to_export[col].dt.tz is not None: df_to_export[col] = df_to_export[col].dt.tz_localize(None)
            df_to_export.to_excel(writer, index=False, sheet_name='ServiceData')
        excel_data = output.getvalue()
        st.sidebar.download_button(label="Download Filtered Data as XLSX", data=excel_data,
            file_name='filtered_service_data.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', key='download_xlsx')

else:
    st.info("No data to display. Please ensure the Google Sheet is accessible, contains data with headers for all expected columns, and 'Credentials.json' is correctly set up.")

